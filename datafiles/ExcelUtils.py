
#Install openpyxl package
import openpyxl


def get_row_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)


def get_column_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)


def read_data(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum, column=columnNum).value


def write_data(file, sheetName, rowNum, columnNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row =rowNum, column=columnNum).value = data
    workbook.save(file)


